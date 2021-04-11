import pymysql
from faker import Faker
from openpyxl import Workbook
import random
import time


# # 循环写入数据库
# conn = pymysql.connect(host="localhost", port=3306, user="root", password="1qaz@WSX", db="test",
#                        charset="utf8")
#
# cursor = conn.cursor()
# f = Faker("zh-CN")
#
# for i in range(100):
#     sql = """insert into test( name,company,job,phone_number,address) values ('%s','%s','%s','%s','%s')""" % (
#         f.name(), f.company(), f.job(), f.phone_number(), f.address())
#
#     cursor.execute(sql)
#
# conn.commit()
# cursor.close()
# conn.close()


# 批量制造写入excel

def write_excel(file_Name, data):
    wb = Workbook()
    # 写入表头
    excel_head = ['人员姓名', '人员编号', '所属组织', '证件类型', '工作电话', '电子邮箱', '排序号',
                  '性别', '出生日期', '人员分类', '证件号码', '手机号码', '人员状态', '生效时间', '失效时间',
                  '账户', '账户生效时间', '组织', '岗位', '职务', '职级', '生效时间', '失效时间',
                  '扩展字段1文本类型', '数字类型字段', '日期类型字段', '时间类型格式', '枚举类型', '选人类型', '选部门类型']
    sheet0Name = '个人信息'

    sheet0 = wb.create_sheet(sheet0Name, index=0)
    sheet0.merge_cells(range_string=None,
                       start_row=1,
                       start_column=1,
                       end_row=1,
                       end_column=30)
    sheet0.cell(row=1, column=1, value='''
    说明:\t
    1，更新数据时，将会根据重复验证字段进行数据更新\t
    2，跳过不处理，将会根据重复验证字段跳过\t
    3，机构将使用机构名称(机构编号），人员将使用人员名称(人员编号），多个使用、隔开\t
    4，名称，编号，所属组织，主岗是必填\t
    5，岗位信息(岗位，职务，职级)\t
    ''')
    sheet0.cell(row=2, column=1, value="基本信息")
    sheet0.cell(row=2, column=16, value="账户信息")
    sheet0.cell(row=2, column=18, value="主岗信息")
    sheet0.cell(row=2, column=24, value="扩展分类1")
    for i, item in enumerate(excel_head):
        sheet0.cell(row=3, column=i + 1, value=item)

    # 写入数据
    for i, item in enumerate(data):
        i = i + 4
        for j, val in enumerate(item):
            sheet0.cell(row=i, column=j + 1, value=val)

    wb.save(file_Name + '.xlsx')


# 生成测试数据
def faker_maker(i):
    f = Faker("zh-CN")
    faker_list = []

    # 人员姓名
    name = f.name()
    # 人员编号
    personnel_number = f.random_int()
    # 所属组织
    organization = f.company()
    # 证件类型
    certificate_type = "身份证"
    # 工作电话
    work_phone = f.phone_number()
    # 电子邮箱
    email = f.ascii_free_email()
    # 性别
    sex = random.choice(['男', '女'])
    # 出生日期
    date_of_birth = f.date_of_birth(tzinfo=None, minimum_age=0, maximum_age=115)
    # 人员分类
    personnel_classification = random.choice(['正式员工', '实习生'])
    # 证件号码
    ID_number = f.ssn(min_age=18, max_age=90)
    # 手机号码
    phone_number = f.phone_number()
    # 人员状态
    personnel_status = random.choice(['ON', 'OFF'])
    # 生效时间
    effective_time = f.date_between(start_date="-30y", end_date="today")
    # 失效时间
    invalid_time = f.date_between(start_date="today", end_date="+30y")
    # 账户
    account = f.credit_card_number()
    # 账户生效时间
    account_effective_time = f.date_between(start_date="-1y", end_date="today")
    # 组织
    organize = organization
    # 岗位
    post = f.job()
    # 职务
    job = random.choice(['技术主管', '销售主管', '人事经理'])
    # 职级
    rank = random.choice(['初级', '中级', '高级'])

    print(f'已生成{i + 1}条数据')

    faker_list.append(name)
    faker_list.append(personnel_number)
    faker_list.append(organization)
    faker_list.append(certificate_type)
    faker_list.append(work_phone)
    faker_list.append(email)
    faker_list.append(str(i + 1))
    faker_list.append(sex)
    faker_list.append(date_of_birth)
    faker_list.append(personnel_classification)
    faker_list.append(ID_number)
    faker_list.append(phone_number)
    faker_list.append(personnel_status)
    faker_list.append(effective_time)
    faker_list.append(invalid_time)
    faker_list.append(account)
    faker_list.append(account_effective_time)
    faker_list.append(organize)
    faker_list.append(post)
    faker_list.append(job)
    faker_list.append(rank)
    return faker_list


# 将数据写入excel
def faker_list_add(num):
    base_list = []
    for i in range(num):
        fake_new_list = faker_maker(i)
        base_list.append(fake_new_list)
    return base_list


if __name__ == "__main__":
    start_time = time.time()
    write_excel(r'F:\ym', faker_list_add(10))
    end_time = time.time()
    total_time = end_time - start_time
    print(f'用时{total_time}秒')
